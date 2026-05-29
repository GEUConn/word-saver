#!/usr/bin/env python3
"""
单词快存工具
选中文字后按 Ctrl+Shift+S，自动翻译并保存到桌面「单词本」文件夹中按日期命名的 Excel 文件。
"""

import keyboard
import pyperclip
import time
import os
import sys
import threading
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from deep_translator import GoogleTranslator
from langdetect import detect, LangDetectException
import eng_to_ipa as ipa
from pypinyin import lazy_pinyin, Style
from plyer import notification
import pystray
from PIL import Image, ImageDraw, ImageFont

# ── 配置 ──────────────────────────────────────────────
# 打包成 exe 时用 exe 所在目录，直接运行 .py 时用脚本所在目录
if getattr(sys, "frozen", False):
    SAVE_FOLDER = Path(sys.executable).parent
else:
    SAVE_FOLDER = Path(__file__).parent
HOTKEY = "f8"
# ─────────────────────────────────────────────────────


def create_tray_icon() -> Image.Image:
    """生成系统托盘图标（蓝底白字 W）"""
    size = 64
    img = Image.new("RGB", (size, size), color="#3A7BD5")
    draw = ImageDraw.Draw(img)
    # 画圆角矩形背景
    draw.rounded_rectangle([4, 4, size - 4, size - 4], radius=12, fill="#3A7BD5")
    # 写字母
    try:
        font = ImageFont.truetype("arial.ttf", 36)
    except Exception:
        font = ImageFont.load_default()
    draw.text((18, 10), "W", fill="white", font=font)
    return img


def get_phonetic(text: str, lang: str) -> str:
    """获取音标：英文返回 IPA，中文返回拼音"""
    try:
        if lang == "en":
            result = ipa.convert(text)
            # eng_to_ipa 无法识别时原样返回，排除无效结果
            if result and result != text and "*" not in result:
                return f"/{result}/"
            return ""
        elif lang in ("zh-cn", "zh", "zh-tw"):
            pinyin = " ".join(lazy_pinyin(text, style=Style.TONE))
            return pinyin
    except Exception:
        pass
    return ""


def detect_lang(text: str) -> str:
    """检测语言，失败时默认英文"""
    try:
        return detect(text)
    except LangDetectException:
        return "en"


def translate(text: str, lang: str) -> str:
    """翻译：中文→英文，其他→中文"""
    if lang in ("zh-cn", "zh", "zh-tw"):
        return GoogleTranslator(source="auto", target="en").translate(text)
    else:
        return GoogleTranslator(source="auto", target="zh-CN").translate(text)


def ensure_header(ws):
    """如果工作表是新建的，写入表头"""
    ws.append(["单词 / 词组", "翻译", "音标"])
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="3A7BD5")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22


def save_to_excel(word: str, translation: str, phonetic: str):
    """追加一行到今天的 Excel 文件"""
    SAVE_FOLDER.mkdir(parents=True, exist_ok=True)
    filename = SAVE_FOLDER / f"{date.today().strftime('%Y-%m-%d')}.xlsx"

    if filename.exists():
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = date.today().strftime("%Y-%m-%d")
        ensure_header(ws)

    row = [word, translation, phonetic]
    ws.append(row)

    # 交替行背景色，方便阅读
    row_idx = ws.max_row
    if row_idx % 2 == 0:
        fill = PatternFill("solid", fgColor="EAF1FB")
        for col in range(1, 4):
            ws.cell(row=row_idx, column=col).fill = fill

    wb.save(filename)


def on_hotkey():
    """热键触发后在后台线程执行，避免阻塞键盘监听"""
    def worker():
        print(f"[热键触发] 正在读取选中文字...")
        # 1. 暂存剪贴板原内容
        try:
            old_clip = pyperclip.paste()
        except Exception:
            old_clip = ""

        # 2. 模拟 Ctrl+C 复制选中文字
        keyboard.send("ctrl+c")
        time.sleep(0.15)

        try:
            word = pyperclip.paste().strip()
        except Exception:
            word = ""

        # 3. 还原剪贴板
        try:
            if old_clip:
                pyperclip.copy(old_clip)
        except Exception:
            pass

        # 4. 校验
        if not word:
            return
        if len(word) > 300:
            notify("单词快存", "选中文字太长（>300字），已跳过")
            return

        # 5. 翻译 + 音标
        try:
            lang = detect_lang(word)
            result = translate(word, lang)
            phonetic = get_phonetic(word, lang)
        except Exception as e:
            notify("单词快存 ✗", f"翻译失败：{str(e)[:60]}")
            return

        # 6. 保存
        try:
            save_to_excel(word, result, phonetic)
            notify("✅ 已保存", f"{word}  →  {result}")
        except Exception as e:
            notify("单词快存 ✗", f"保存失败：{str(e)[:60]}")

    threading.Thread(target=worker, daemon=True).start()


def notify(title: str, message: str):
    try:
        notification.notify(title=title, message=message, timeout=2)
    except Exception:
        pass  # 无法弹通知时静默失败


def open_folder(_icon=None, _item=None):
    try:
        os.startfile(str(SAVE_FOLDER))
    except Exception:
        pass


def main():
    print(f"[启动] 单词快存正在运行，热键: {HOTKEY}")
    print(f"[存储] Excel 将保存到: {SAVE_FOLDER}")

    # 注册热键
    try:
        keyboard.add_hotkey(HOTKEY, on_hotkey)
        print(f"[热键] {HOTKEY} 注册成功")
    except Exception as e:
        print(f"[错误] 热键注册失败: {e}")
        print("请尝试以管理员身份运行此程序")
        input("按回车退出...")
        return

    # 系统托盘
    menu = pystray.Menu(
        pystray.MenuItem("打开单词本文件夹", open_folder),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("退出", lambda icon, item: (icon.stop(), os._exit(0))),
    )
    icon = pystray.Icon(
        name="单词快存",
        icon=create_tray_icon(),
        title=f"单词快存  ({HOTKEY.upper()} 保存选中文字)",
        menu=menu,
    )

    notify("单词快存已启动", f"按 {HOTKEY.upper()} 保存选中文字")
    print("[就绪] 程序已启动，选中文字后按热键即可保存")
    icon.run()  # 阻塞，直到用户点「退出」


if __name__ == "__main__":
    main()
